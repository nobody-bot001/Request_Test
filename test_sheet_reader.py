import pandas as pd
from datetime import datetime, timedelta, timezone
import json
import io
import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import tempfile
import traceback
import subprocess
import os

# ---------------------------------------------------------------------
# STEP 1: SAFE HELPERS
# ---------------------------------------------------------------------
def safe_serialize(obj):
    if obj is None:
        return None
    elif isinstance(obj, (str, int, float, bool)):
        return obj
    elif hasattr(obj, 'rgb') and obj.rgb:
        return str(obj.rgb)
    elif hasattr(obj, '__str__'):
        return str(obj)
    else:
        return None

def extract_color(color_obj):
    if not color_obj:
        return None
    if isinstance(color_obj, str):
        return color_obj
    if hasattr(color_obj, 'rgb') and color_obj.rgb:
        return str(color_obj.rgb)
    return str(color_obj)

def argb_to_rgba(argb_color):
    if not argb_color or not isinstance(argb_color, str):
        return argb_color
    clean_color = ''.join(c for c in argb_color if c in '0123456789ABCDEFabcdef')
    if len(clean_color) == 8:
        if clean_color.startswith(('FF', 'ff')):
            return f"#{clean_color[2:]}".upper()
        else:
            return f"#{clean_color[2:]}{clean_color[:2]}".upper()
    elif len(clean_color) == 6:
        return f"#{clean_color}".upper()
    else:
        return argb_color

def get_cell_styles(cell):
    if not cell:
        return {
            "font": {"name": None, "size": None, "bold": None, "italic": None, "color": None},
            "fill": {"fgColor": None},
            "alignment": {"horizontal": None, "vertical": None, "wrap_text": None}
        }
    font_color = None
    if cell.font and cell.font.color:
        raw_color = extract_color(cell.font.color)
        font_color = argb_to_rgba(raw_color)
    fill_color = None
    if cell.fill and cell.fill.fgColor:
        raw_color = extract_color(cell.fill.fgColor)
        fill_color = argb_to_rgba(raw_color)
    return {
        "font": {
            "name": safe_serialize(cell.font.name) if cell.font else None,
            "size": safe_serialize(cell.font.size) if cell.font else None,
            "bold": safe_serialize(cell.font.bold) if cell.font else None,
            "italic": safe_serialize(cell.font.italic) if cell.font else None,
            "color": font_color
        },
        "fill": {"fgColor": fill_color},
        "alignment": {
            "horizontal": safe_serialize(cell.alignment.horizontal) if cell.alignment else None,
            "vertical": safe_serialize(cell.alignment.vertical) if cell.alignment else None,
            "wrap_text": safe_serialize(cell.alignment.wrap_text) if cell.alignment else None
        }
    }

class SafeJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if hasattr(obj, '__str__'):
            return str(obj)
        return super().default(obj)

# ---------------------------------------------------------------------
# STEP 2: DOWNLOAD GOOGLE SHEET
# ---------------------------------------------------------------------
try:
    # Previous Sheet (commented)
    # SHEET_ID = "1PttAjbFtfvymn9pBpvZdMwNylzQNFVC9yGzvsMl2Vr8"
    # GID = "0"

    SHEET_ID = "1cmDXt7UTIKBVXBHhtZ0E4qMnJrRoexl2GmDFfTBl0Z4"
    GID = "1882612924"
    csv_url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx&gid={GID}"

    print("📥 Fetching Google Sheet as XLSX...")
    r = requests.get(csv_url)
    r.raise_for_status()

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    temp_file.write(r.content)
    temp_file.close()
    print("✅ Sheet downloaded and stored temporarily.")

    df_preview = pd.read_excel(temp_file.name)
    print("✅ Loaded Google Sheet Preview:\n", df_preview.head())

    utc_time = datetime.now(timezone.utc)
    local_time = utc_time + timedelta(hours=5)

    status_info = {
        "status": "success",
        "message": "Google Sheet downloaded successfully",
        "last_updated": utc_time.strftime("%Y-%m-%d %H:%M:%S UTC"),
        "local_time": local_time.strftime("%Y-%m-%d %I:%M:%S %p PKT"),
        "sheet_id": SHEET_ID,
        "gid": GID
    }

    with open("success.json", "w", encoding="utf-8") as f:
        json.dump(status_info, f, ensure_ascii=False, indent=4)

except Exception as e:
    error_info = {
        "status": "error",
        "message": str(e),
        "traceback": traceback.format_exc()
    }
    with open("success.json", "w", encoding="utf-8") as f:
        json.dump(error_info, f, ensure_ascii=False, indent=4)
    print("❌ Error downloading Google Sheet:", e)
    raise SystemExit()

# ---------------------------------------------------------------------
# STEP 3: EXTRACT STRUCTURED TIMETABLE
# ---------------------------------------------------------------------
try:
    wb = load_workbook(temp_file.name, data_only=True)
    structured_timetable = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        structured_timetable[sheet_name] = {"rooms": {}, "time_slots": [], "all_cells": []}

        header_rows = []
        for row in range(1, sheet.max_row + 1):
            first_cell = sheet.cell(row=row, column=1).value
            if first_cell and str(first_cell).strip().lower() in ["room", "rooms", "lab", "labs", "timeslots", "time slots"]:
                header_rows.append(row)
        if 5 not in header_rows:
            header_rows.insert(0, 5)

        time_slot_maps = {}
        for hr in header_rows:
            slot_map = {}
            for col in range(2, sheet.max_column + 1):
                cell = sheet.cell(row=hr, column=col)
                if cell.value:
                    slot_map[col] = safe_serialize(cell.value)
            time_slot_maps[hr] = slot_map

        for slot_map in time_slot_maps.values():
            structured_timetable[sheet_name]["time_slots"].extend(slot_map.values())

        rooms = {}
        for row in range(6, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=1)
            if cell.value:
                room_name = safe_serialize(cell.value)
                if room_name:
                    rooms[row] = room_name
                    structured_timetable[sheet_name]["rooms"][room_name] = {
                        "time_slots": {},
                        "row": row,
                        "schedule": []
                    }

        for row in range(1, sheet.max_row + 1):
            header_row = max([hr for hr in header_rows if hr <= row], default=5)
            time_slots = time_slot_maps[header_row]
            row_data = []
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell_info = {
                    "cell_reference": f"{get_column_letter(col)}{row}",
                    "row": row,
                    "column": col,
                    "column_letter": get_column_letter(col),
                    "value": safe_serialize(cell.value),
                    "time_slot": time_slots.get(col),
                    "room": rooms.get(row),
                    "styles": get_cell_styles(cell)
                }
                row_data.append(cell_info)

                if row > header_row and col > 1 and cell.value:
                    room_name = rooms.get(row)
                    time_slot = time_slots.get(col)
                    subject = safe_serialize(cell.value)
                    if room_name and subject:
                        if time_slot not in structured_timetable[sheet_name]["rooms"][room_name]["time_slots"]:
                            structured_timetable[sheet_name]["rooms"][room_name]["time_slots"][time_slot] = []
                        class_info = {
                            "subject": subject,
                            "time_slot": time_slot,
                            "cell_reference": f"{get_column_letter(col)}{row}",
                            "row": row,
                            "column": col,
                            "styles": get_cell_styles(cell)
                        }
                        structured_timetable[sheet_name]["rooms"][room_name]["time_slots"][time_slot].append(class_info)
                        structured_timetable[sheet_name]["rooms"][room_name]["schedule"].append(class_info)
            structured_timetable[sheet_name]["all_cells"].append(row_data)

    with open("timetable_detailed.json", "w", encoding="utf-8") as f:
        json.dump(structured_timetable, f, indent=2, ensure_ascii=False, cls=SafeJSONEncoder)
    print("✅ Created detailed structured timetable (timetable_detailed.json)")

    simplified_timetable = {}
    for sheet_name, sheet_data in structured_timetable.items():
        simplified_timetable[sheet_name] = []
        color_groups = {}
        if sheet_data["all_cells"]:
            max_cols = len(sheet_data["all_cells"][0])
            for row in range(1, min(5, len(sheet_data["all_cells"]) + 1)):
                for col in range(7, max_cols + 1):
                    cell = sheet_data["all_cells"][row - 1][col - 1]
                    if cell["value"] and cell["styles"]["fill"]["fgColor"]:
                        color_groups[cell["styles"]["fill"]["fgColor"]] = cell["value"]

        for room_name, room_data in sheet_data["rooms"].items():
            room_schedule = {"room": room_name, "schedule": []}
            for class_info in room_data["schedule"]:
                fg_color = class_info["styles"]["fill"]["fgColor"]
                linked_group = color_groups.get(fg_color)
                schedule_entry = {
                    "time_slot": class_info["time_slot"],
                    "subject": class_info["subject"],
                    "location": f"Cell {class_info['cell_reference']}",
                    "fgColor": fg_color,
                    "Color_Linked": linked_group,
                    "row": class_info["row"],
                    "column": class_info["column"]
                }
                room_schedule["schedule"].append(schedule_entry)
            room_schedule["schedule"].sort(key=lambda x: x["column"])
            simplified_timetable[sheet_name].append(room_schedule)

    with open("timetable_simplified.json", "w", encoding="utf-8") as f:
        json.dump(simplified_timetable, f, indent=2, ensure_ascii=False, cls=SafeJSONEncoder)
    print("✅ Created simplified timetable (timetable_simplified.json)")
    print("🎉 Timetable extraction completed successfully!")

except Exception as e:
    print("❌ Error during timetable extraction:", e)
    error_info = {
        "status": "error",
        "message": str(e),
        "traceback": traceback.format_exc()
    }
    with open("success.json", "w", encoding="utf-8") as f:
        json.dump(error_info, f, ensure_ascii=False, indent=4)
    raise SystemExit()

# ---------------------------------------------------------------------
# STEP 4: PUSH TO GITHUB
# ---------------------------------------------------------------------
try:
    print("🚀 Pushing JSON files to GitHub...")

    subprocess.run(["git", "config", "--global", "user.name", "github-actions[bot]"])
    subprocess.run(["git", "config", "--global", "user.email", "github-actions[bot]@users.noreply.github.com"])

    subprocess.run(["git", "add", "timetable_detailed.json"])
    subprocess.run(["git", "add", "timetable_simplified.json"])
    subprocess.run(["git", "add", "success.json"])
    subprocess.run(["git", "commit", "-m", "Updated timetable JSON files automatically"])
    subprocess.run(["git", "push"])

    print("✅ Successfully pushed all JSON files to GitHub.")

except Exception as e:
    print("❌ GitHub push failed:", e)
    error_info = {
        "status": "error",
        "message": "GitHub push failed",
        "error_detail": str(e),
        "traceback": traceback.format_exc()
    }
    with open("success.json", "w", encoding="utf-8") as f:
        json.dump(error_info, f, ensure_ascii=False, indent=4)



# import pandas as pd
# from datetime import datetime, timedelta, timezone
# import json
# import io
# import requests

# # Previous Google Sheet (commented out)
# # SHEET_ID = "1PttAjbFtfvymn9pBpvZdMwNylzQNFVC9yGzvsMl2Vr8"
# # GID = "0"

# # New Google Sheet
# SHEET_ID = "1cmDXt7UTIKBVXBHhtZ0E4qMnJrRoexl2GmDFfTBl0Z4"
# GID = "1882612924"

# csv_url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/gviz/tq?tqx=out:csv&gid={GID}"

# try:
#     r = requests.get(csv_url)
#     r.raise_for_status()

#     df = pd.read_csv(io.StringIO(r.text))
#     print("✅ Loaded Google Sheet data:\n", df)

#     data = df.to_dict(orient="records")

#     utc_time = datetime.now(timezone.utc)
#     local_time = utc_time + timedelta(hours=5)  # Pakistan Standard Time (UTC+5)

#     result = {
#         "last_updated": utc_time.strftime("%Y-%m-%d %H:%M:%S UTC"),
#         "local_time": local_time.strftime("%Y-%m-%d %I:%M:%S %p PKT"),
#         "data": data
#     }

#     with open("sheet_backup.json", "w", encoding="utf-8") as f:
#         json.dump(result, f, ensure_ascii=False, indent=4)

#     print("💾 Saved as sheet_backup.json with local time included")

# except Exception as e:
#     print("❌ Error:", e)

